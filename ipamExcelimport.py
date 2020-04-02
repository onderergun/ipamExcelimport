
import argparse
import json
import time
import requests
from getpass import getpass
from urllib3.exceptions import InsecureRequestWarning

from openpyxl import load_workbook

requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

def createPool(cvpServer, session_id, session_token, name, description, parentid, subnet, vlan, gateway, notificationemails):
    url_params = {'name':name, 'description':description, 'parentid': parentid, 'notificationemails': notificationemails, 'subnet':subnet, 'vlan':vlan, 'gateway':gateway, 'pingsweep':True, 'dnslookup':True, 'emailwarning':75, 'emailcritical':90}
    update_resp = requests.post(
        'https://%s/cvp-ipam-api/pool?session_id=%s&token=%s'% (cvpServer, session_id, session_token),data=json.dumps(url_params), verify=False)
    return update_resp.json()

def createReservation(cvpServer, session_id, session_token, description, parentid, range):
    url_params = { 'description':description, 'parentid': parentid, 'range': range, 'group':description}
    update_resp = requests.post(
        'https://%s/cvp-ipam-api/reservation?session_id=%s&token=%s'% (cvpServer, session_id, session_token),data=json.dumps(url_params), verify=False)
    return update_resp.json()

def createAllocation(cvpServer, session_id, session_token, value, description, parentid, subnet):
    url_params = {'value':value, 'description':description, 'parentid': parentid, 'subnet': subnet}
    update_resp = requests.post(
        'https://%s/cvp-ipam-api/allocation?session_id=%s&token=%s'% (cvpServer, session_id, session_token),data=json.dumps(url_params), verify=False)
    return update_resp.json()

def main():
    
    parser = argparse.ArgumentParser()
    parser.add_argument('--username', required=True)
    parser.add_argument('--cvpServer', required=True)
    parser.add_argument('--notificationemails', required=True)
    parser.add_argument('--filename', required=True)
    
    args = parser.parse_args()
    username = args.username
    cvpServer=args.cvpServer
    notificationemails = args.notificationemails
    password = getpass()
    filename =  args.filename
    
    print ('Start Login')
    login_data = {'username': username, 'password': password}
    login_resp = requests.post('https://%s/cvp-ipam-api/login' % cvpServer,
                               data=json.dumps(login_data), verify=False)
    print ('Login Info')
    login_json = login_resp.json()
    print ('\n')

    session_id = login_json['session_id']
    session_token = login_json['token']
    
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        octets= sheet.split(".")
        subnet = ".".join(octets[0:3]) + ".0/24"
        currentsheet = wb[sheet]
        name = currentsheet['D2'].value
        description = name
        vlan = str(currentsheet['B2'].value)
        gateway = ".".join(octets[0:3]) + ".254"
        parentid = "network1-ipv4"
        
        # Creating subnet
        print ("Creating Subnet ", subnet)
        print (name, description, parentid, subnet, vlan, gateway, notificationemails)
        response = createPool(cvpServer, session_id, session_token, name, description, parentid, subnet, vlan, gateway, notificationemails)
        if response['success']:
            print ("Success")

        # Creating reservation
        subnetstart = ".".join(octets[0:3]) + ".1"
        subnetend = ".".join(octets[0:3]) + ".253"
        range = subnetstart + "|" + subnetend
        print ("Creating Reservation Range ", range)
        parentid = parentid + "-" + name
        response = createReservation(cvpServer, session_id, session_token, description, parentid, range)
        if response['success']:
            print ("Success")
        
        # Allocating IP Addresses
        row_count = currentsheet.max_row
        row = 2
        parentid = parentid + "-" + range
        subnet = ""
        while row < row_count:
            if (currentsheet['C'+str(row)].value != None):
                print (currentsheet['A'+str(row)].value, currentsheet['C'+str(row)].value)
                value = currentsheet['A'+str(row)].value
                description = currentsheet['C'+str(row)].value
                response = createAllocation(cvpServer, session_id, session_token, value, description, parentid, subnet)
                if response['success']:
                    print ("Success")
            row +=1
    
    print ('\n')
    print ('Start Logout')
    logout_data = {'session_id': session_id, 'token': session_token}
    logout_resp = requests.post('https://%s/cvp-ipam-api/logout' % cvpServer,
                                data=json.dumps(logout_data), verify=False)
    print ('Logout Info')
    logout_json = logout_resp.json()
    print (logout_json)



if __name__ == '__main__':
    main()


